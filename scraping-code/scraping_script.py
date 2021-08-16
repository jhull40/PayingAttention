### INPUTS
PATH_TO_BROWSER = "C:/chromedriver.exe"
INPUT_FILE = "INPUT.xlsx"
INPUT_SHEET = "Sheet1"
USERNAME = "testaccnba"
PASSWORD = "s8d9jasd90ad2" # registered to email: nolib71514@topmail1.net
DATABASE_NAME = "ODDSPORTAL_DATABASE.db"
### END OF INPUTS


# install: selenium, lxml, openpyxl

from selenium import webdriver  #install selenium
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import time
import pprint
import os
import sys
from datetime import datetime
import json
from lxml import html # pip install lxml
from openpyxl import Workbook # pip install openpyxl
from openpyxl import load_workbook
import csv
import sqlite3



TIME_PAUSE = 1.0 # pause

def wait_by_xpath(xp, how_long_to_wait): # xp is string, how_long_to_wait float - the number of seconds to wait
    try:
        WebDriverWait(driver, how_long_to_wait).until(EC.presence_of_element_located((By.XPATH, xp)) )
        time.sleep(TIME_PAUSE)
        return 1 # success
    except TimeoutException:
        print ("Too much time has passed while waiting for", xp)
        return 0 # fail

        
def fix_string(entry_string): # remove "\n", "\t" and double spaces
    exit_string = entry_string.replace("\n", "")
    exit_string = exit_string.replace("\t", "")
    exit_string = exit_string.replace("\r", "")
    while "  " in exit_string:
        exit_string = exit_string.replace("  ", " ")
    if len(exit_string) > 0: # remove first space
        if exit_string[0] == ' ':
            exit_string = exit_string[1:len(exit_string)]
    if len(exit_string) > 0: # remove last space
        if exit_string[len(exit_string)-1] == ' ':
            exit_string = exit_string[0:len(exit_string)-1]

    return exit_string


def start_driver_normal():
    normal_driver = webdriver.Chrome(PATH_TO_BROWSER)
    normal_driver.maximize_window()
    return normal_driver


def start_driver_headless():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("window-size=1920,1080")
    headless_driver = webdriver.Chrome(PATH_TO_BROWSER, options=chrome_options)
    return headless_driver





def destroy_handles_and_create_new_one():
    # call it before opening an url
    while 1:
        initial_handles = driver.window_handles
        driver.execute_script("window.open();")
        handles_after_opening = driver.window_handles
        if len(handles_after_opening) > len(initial_handles):
            break
        else:
            print("Couldn't open a handle!")
            time.sleep(10.0)
            continue
        
    added_handle = []
    for handle in handles_after_opening:
        if handle in initial_handles:
            driver.switch_to.window(handle)
            driver.close()
        else:
            added_handle.append(handle)

    driver.switch_to.window(added_handle[0])
    return


def read_inputs():
    inputs_to_return = []

    if not os.path.exists(INPUT_FILE):
        print("Input file", INPUT_FILE, "does not exist!")
        return inputs_to_return # return empty

    # file does exist, see if the sheet exists
    input_wb = load_workbook(INPUT_FILE)
    try:
        input_ws = input_wb[INPUT_SHEET]
    except KeyError:
        print("There is no sheet", INPUT_SHEET, "in file", INPUT_FILE)
        return inputs_to_return # return empty

    # if here, it should be good to read
    for input_row_index in range(2, input_ws.max_row+1):
        potential_link = input_ws.cell(row=input_row_index, column=1).value
        cur_season_param_input = input_ws.cell(row=input_row_index, column=2).value
        if type(potential_link) !=  str or type(cur_season_param_input) != str:
            print("Row", input_row_index, "in input file doesn't have a valid type value in either column 1 or 2.")
            continue

        # check link validity and yes/no validity
        cur_season_param_input = cur_season_param_input.lower()
        if cur_season_param_input not in ["yes", "no"]:
            print("Row", input_row_index, "should have a yes/no value in column B.")
            continue

        if "www.oddsportal.com" not in potential_link:
            print("Row", input_row_index, "doesn't have a oddsportal link in column A.")
            continue

        # if here, should be good
        if cur_season_param_input == 'yes':
            cur_season_param = True
        else:
            cur_season_param = False
        input_dict_to_add = {"season_link": potential_link, "is_current": cur_season_param}
        if input_dict_to_add not in inputs_to_return:
            inputs_to_return.append(input_dict_to_add)
   
    return inputs_to_return


def login():
    while 1:
        try:
            driver.get("https://www.oddsportal.com/login/")
            wait_for_login = wait_by_xpath("//input[@id='login-username1']", 20)

            logout_el = driver.find_elements_by_xpath("//li[@id='user-header-logout']")
            if len(logout_el) != 0:
                print("Logged in successfully.")
                return

            driver.find_element_by_xpath("//input[@id='login-username1']").clear()
            driver.find_element_by_xpath("//input[@id='login-username1']").send_keys(USERNAME)
            driver.find_element_by_xpath("//input[@id='login-password1']").send_keys(PASSWORD)
            driver.find_element_by_xpath("//input[@id='login-password1']").send_keys(Keys.RETURN)

            # wait for actualization
            time.sleep(10.0)
            logout_el = driver.find_elements_by_xpath("//li[@id='user-header-logout']")
            if len(logout_el) != 0:
                print("Logged in successfully.")
                return
        except:
            print("Some kind of exception happened while logging in!")
            continue
            
        
if __name__ == '__main__':
    ## read input items
    season_links = read_inputs()

    ## create database
    db_conn = sqlite3.connect(DATABASE_NAME)
    db_cursor = db_conn.cursor()
    db_cursor.execute("CREATE TABLE IF NOT EXISTS FullyScrapedSeasonLinks (season_link TEXT NOT NULL PRIMARY KEY)")
    db_cursor.execute("""CREATE TABLE IF NOT EXISTS GamesTable (game_link TEXT NOT NULL PRIMARY KEY, home_team TEXT, away_team TEXT, score_home INTEGER, score_away INTEGER,
                    date_string TEXT, hour_string TEXT, date_unix REAL, full_scoreline TEXT, odds_json TEXT)""")


    ## start chrome, login on the site
    driver = start_driver_headless()
    login() # won't complete until it logins!
    
    ## go through each season, check if it is already fully scraped!
    print("Getting a list of games...")
    for season_item in season_links:
        # check if season is fully scraped, only check for non current seasons.
        if season_item["is_current"] == False:
            existence_check = db_cursor.execute("SELECT EXISTS(SELECT 1 FROM FullyScrapedSeasonLinks WHERE season_link=?)", (season_item["season_link"],)).fetchone()[0]
            if existence_check == 1:
                print("Season URL", season_item["season_link"], "was already scraped!")
                continue

        # if here, must scrape this season! for each game, get: full link, home team, away team, home score and away score (no OT, just pure numbers)
        this_season_games = {}
        current_page = 1
        while 1:
            if current_page == 1:
                page_url_to_open = season_item["season_link"]
            else:
                page_url_to_open = next_page_url

            try:
                destroy_handles_and_create_new_one()
                driver.get(page_url_to_open)
                wait_for_games = wait_by_xpath("//table[@id='tournamentTable']/tbody/tr/td[@class='name table-participant']/..", 20)
                if wait_for_games == 0:
                    continue
                innerHTML_gamelist = driver.execute_script("return document.body.innerHTML")
                htmlElem_gamelist = html.document_fromstring(innerHTML_gamelist)

            except KeyboardInterrupt:
                print("Manual interrupt, quit!")
                driver.quit()
                db_cursor.close()
                db_conn.close()
                sys.exit(0)
            except:
                print("An exception while opening", page_url_to_open)
                continue

            # get games, get next page url.
            number_of_games_found_on_this_page = 0 # has to be different than 0 to work.
            games_els = htmlElem_gamelist.xpath("//table[@id='tournamentTable']/tbody/tr/td[@class='name table-participant']/..")
            for game_el in games_els:
                game_to_add = {}

                game_to_add["url"] = ''
                game_to_add["home"] = ''
                game_to_add["away"] = ''
                game_to_add["homescore"] = ''
                game_to_add["awayscore"] = ''

                url_el = game_el.xpath("./td[@class='name table-participant']/a[@href]")
                if len(url_el) != 0:
                    game_to_add["url"] = 'https://www.oddsportal.com' + url_el[0].attrib["href"]

                teams_els = game_el.xpath("./td[@class='name table-participant']/a[@href]/node()")
                for team_el in teams_els:
                    try:
                        team_text = team_el.text_content()
                        team_text = fix_string(team_text)
                    except AttributeError:
                        team_text = fix_string(team_el)

                    if team_text != '':
                        # good to add, depends which one!
                        team_text = fix_string(team_text.replace("-", ""))
                        if game_to_add["home"] == '':
                            game_to_add["home"] = team_text
                        elif game_to_add["away"] == '':
                            game_to_add["away"] = team_text
                        else:
                            continue

                score_el = game_el.xpath("./td[@class='center bold table-odds table-score']")
                if len(score_el) != 0:
                    both_scores_text = score_el[0].text_content()
                    both_scores_array = both_scores_text.split(":")
                    if len(both_scores_array) == 2:
                        game_to_add["homescore"] = "".join(char for char in both_scores_array[0] if char.isdigit())
                        game_to_add["awayscore"] = "".join(char for char in both_scores_array[1] if char.isdigit())
                        if game_to_add["homescore"] != '':
                            game_to_add["homescore"] = int(game_to_add["homescore"])
                        if game_to_add["awayscore"] != '':
                            game_to_add["awayscore"] = int(game_to_add["awayscore"])

                # check validity
                if game_to_add["url"] != '' and game_to_add["home"] != '' and game_to_add["away"] != '' and game_to_add["homescore"] != '' and game_to_add["awayscore"] != '':
                    number_of_games_found_on_this_page = number_of_games_found_on_this_page +1
                    if game_to_add["url"] not in this_season_games:
                        this_season_games[game_to_add["url"]] = game_to_add

            # check if anything got scraped!
            if number_of_games_found_on_this_page == 0:
                print("Can't parse any games on URL", page_url_to_open, "something could be wrong!")
                continue

            # find next page
            next_page_el = htmlElem_gamelist.xpath("//div[@id='pagination']/span[@class='active-page']/following-sibling::a[@href][1]/span[not(@class)]/..")
            if len(next_page_el) == 0:
                break
            else:
                current_page = current_page +1
                next_page_url = season_item["season_link"] + next_page_el[0].attrib["href"]


        ## done with pagination, now just try to save all games and full pagination link if needed
        games_preinsertion = db_cursor.execute("SELECT COUNT(*) FROM GamesTable").fetchone()[0]
        for game_to_insert in this_season_games:
            db_cursor.execute("INSERT OR IGNORE INTO GamesTable (game_link, home_team, away_team, score_home, score_away) VALUES(?,?,?,?,?)",
                              (this_season_games[game_to_insert]["url"], this_season_games[game_to_insert]["home"], this_season_games[game_to_insert]["away"],
                               this_season_games[game_to_insert]["homescore"], this_season_games[game_to_insert]["awayscore"] ))
        if season_item["is_current"] == False:
            db_cursor.execute("INSERT OR IGNORE INTO FullyScrapedSeasonLinks (season_link) VALUES(?)", (season_item["season_link"],))
        db_conn.commit()
        games_postinsertion = db_cursor.execute("SELECT COUNT(*) FROM GamesTable").fetchone()[0]
        print("Inserted", games_postinsertion-games_preinsertion, "new games for season", season_item["season_link"])
        







    ## go into each game where odds are null, scrape it! check for being logged in while scraping
    games_to_scrape = db_cursor.execute("SELECT game_link, rowid FROM GamesTable WHERE odds_json IS NULL")
    for game_to_scrape in games_to_scrape.fetchall():
        try:
            destroy_handles_and_create_new_one()
            driver.get(game_to_scrape[0])
            wait_for_gamedata = wait_by_xpath("//div[@id='tab-nav-main']/div[@id='bettype-tabs-scope']/ul[@class='sub-menu subactive']/li[contains(@class, 'active')]/strong", 10)
            if wait_for_gamedata == 0:
                continue
            innerHTML = driver.execute_script("return document.body.innerHTML")
            htmlElem = html.document_fromstring(innerHTML)

            ## verify that script is logged in, then take data that needs to be taken.
            ## essential data: date string, hour string, date unix, odds data
            ## non-essential data: full scoreline, score type, average odds, highest odds, time of scraping.
            is_logged_in_el = htmlElem.xpath("//li[@id='user-header-logout']")
            if len(is_logged_in_el) == 0:
                login()
                continue

            # if here, script is logged in.
            date_string_to_save = '' # essential
            hour_string_to_save = '' # essential
            unix_time_to_save = None # essential
            full_scoreline_to_save = '' # not essential
            odds_data_to_save = {'time_of_scraping':datetime.now().strftime("%m-%d-%Y %H:%M:%S"), 'odds':[], "average_home":'', "average_away":'',
                                 "highest_home":'', "highest_away":"", 'odds_type':'', 'sport':''} # essential, have avg and max odds and time of scraping
            
            time_el = htmlElem.xpath("//div[@id='col-content']/p[contains(@class, 'date datet')]")
            if len(time_el) != 0:
                full_time_string = fix_string(time_el[0].text_content())
                full_time_array = [fix_string(time_item) for time_item in full_time_string.split(",")] # index 2 should be date string and index 3 should be hour string
                if len(full_time_array) == 3:
                    try:
                        datetime_game_object = datetime.strptime(full_time_array[1] + " " + full_time_array[2], "%d %b %Y %H:%M")
                        date_string_to_save = full_time_array[1]
                        hour_string_to_save = full_time_array[2]
                        unix_time_to_save = datetime_game_object.timestamp()
                    except ValueError:
                        print("Can't parse time properly at URL", game_to_scrape[0])

            scoreline_els = htmlElem.xpath("//div[@id='col-content']/div[@id='event-status']/p[@class='result']/span[@class]/following-sibling::node()")
            for scoreline_el in scoreline_els:
                try:
                    scoreline_textpart = scoreline_el.text_content()
                    scoreline_textpart = fix_string(scoreline_textpart)
                except AttributeError:
                    scoreline_textpart = fix_string(scoreline_el)

                if scoreline_textpart != '':
                    full_scoreline_to_save = full_scoreline_to_save + scoreline_textpart + " "
            full_scoreline_to_save = fix_string(full_scoreline_to_save)

            # get odds
            bookies_els = htmlElem.xpath("//table[contains(@class, 'table-main detail-odds')]/tbody/tr/td[1]/div/a[@class='name']/../../..")
            for bookie_el in bookies_els:
                # find name, home and away
                bookie_name_el = bookie_el.xpath("./td[1]/div/a[@class='name']")
                bookie_home_el = bookie_el.xpath("./td[2]")
                bookie_away_el = bookie_el.xpath("./td[3]")

                if len(bookie_name_el) != 0 and len(bookie_home_el) != 0 and len(bookie_away_el) != 0:
                    bookie_name = bookie_name_el[0].text_content()
                    # probe odds
                    try:
                        probe_home_odds = float(bookie_home_el[0].text_content())
                        probe_away_odds = float(bookie_away_el[0].text_content())
                        odds_data_to_save["odds"].append({"bookie": bookie_name, "home":bookie_home_el[0].text_content(), "away": bookie_away_el[0].text_content()})
                    except ValueError:
                        continue

            # get average odds
            avg_home_el = htmlElem.xpath("//table[contains(@class, 'table-main detail-odds')]/tfoot/tr[@class='aver']/td[2]")
            if len(avg_home_el) != 0:
                try:
                    probe_avg_odds = float(avg_home_el[0].text_content())
                    odds_data_to_save["average_home"] = avg_home_el[0].text_content()
                except ValueError:
                    pass
            avg_away_el = htmlElem.xpath("//table[contains(@class, 'table-main detail-odds')]/tfoot/tr[@class='aver']/td[3]")
            if len(avg_away_el) != 0:
                try:
                    probe_avg_away_odds = float(avg_away_el[0].text_content())
                    odds_data_to_save["average_away"] = avg_away_el[0].text_content()
                except ValueError:
                    pass

            # get highest odds
            high_home_el = htmlElem.xpath("//table[contains(@class, 'table-main detail-odds')]/tfoot/tr[@class='highest']/td[2]")
            if len(high_home_el) != 0:
                try:
                    probe_high_home = float(high_home_el[0].text_content())
                    odds_data_to_save["highest_home"] = high_home_el[0].text_content()
                except ValueError:
                    pass
            high_away_el = htmlElem.xpath("//table[contains(@class, 'table-main detail-odds')]/tfoot/tr[@class='highest']/td[3]")
            if len(high_away_el) != 0:
                try:
                    probe_high_away = float(high_away_el[0].text_content())
                    odds_data_to_save["highest_away"] = high_away_el[0].text_content()
                except ValueError:
                    pass

            # get odds type
            odds_type_el = htmlElem.xpath("//div[@id='tab-nav-main']/div[@id='bettype-tabs-scope']/ul[@class='sub-menu subactive']/li[contains(@class, 'active')]/strong")
            if len(odds_type_el) != 0:
                odds_data_to_save["odds_type"] = fix_string(odds_type_el[0].text_content().replace("\xa0", " "))

            # get sport type
            sport_type_el = htmlElem.xpath("//div[@id='main']/div[@id='breadcrumb']/a[@href]")
            if len(sport_type_el) != 0:
                odds_data_to_save["sport"] = sport_type_el[-1].text_content()


            ## check if everything is good, if yes then save
            if date_string_to_save != '' and hour_string_to_save != '' and unix_time_to_save != None and len(odds_data_to_save["odds"])>0:
                # it is good
                db_cursor.execute("UPDATE GamesTable SET date_string=?, hour_string=?, date_unix=?, full_scoreline=?, odds_json=? WHERE rowid=?",
                                  (date_string_to_save, hour_string_to_save, unix_time_to_save, full_scoreline_to_save, json.dumps(odds_data_to_save), game_to_scrape[1]))
                db_conn.commit()
                print("Successfully scraped game in row", game_to_scrape[1])
            else:
                # not good
                print("Something can't be parsed correctly at URL", game_to_scrape[0])
            
 
                    
        except KeyboardInterrupt:
            print("Manual interrupt, quit!")
            driver.quit()
            db_cursor.close()
            db_conn.close()
            sys.exit(0)
        except:
            print("An exception at URL", game_to_scrape[0])
            continue

        
            
    ## end of program
    driver.quit()
    db_cursor.close()
    db_conn.close()
